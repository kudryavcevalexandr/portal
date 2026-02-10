import os
import json
import requests
from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
from copy import copy
from flask_cors import CORS
import tempfile
from datetime import datetime, timezone
from queue import Queue, Full

import psycopg2
import psycopg2.extras
import subprocess
import threading

OPENSEARCH_URL = os.getenv("OPENSEARCH_URL", "http://opensearch:9200").rstrip("/")
DEFAULT_INDEX  = os.getenv("OPENSEARCH_INDEX", "class_tree_v1")
SEARCH_FIELDS  = os.getenv(
    "SEARCH_FIELDS",
    "l4_name^3,l3_name^2,l2_name^2,l1_name,path_name^2,l4_code,path_code"
).split(",")

TIMEOUT = float(os.getenv("HTTP_TIMEOUT", "10"))
XLSX_TEMPLATE = os.getenv("XLSX_TEMPLATE", "/templates/vom_template.xlsx")
XLSX_SHEET    = os.getenv("XLSX_SHEET", "Каталог ВОМа")

# Postgres (Directus DB)
PG_HOST = os.getenv("PG_HOST", "pg")
PG_PORT = int(os.getenv("PG_PORT", "5432"))
PG_DB   = os.getenv("PG_DB", "directus")
PG_USER = os.getenv("PG_USER", "directus")
PG_PASS = os.getenv("PG_PASS", "directus")


app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

ACTION_LOG_FILE = os.getenv("ACTION_LOG_FILE", "/app/logs/actions.log")
ACTION_LOG_QUEUE_SIZE = int(os.getenv("ACTION_LOG_QUEUE_SIZE", "10000"))

action_log_queue = Queue(maxsize=ACTION_LOG_QUEUE_SIZE)


def ensure_action_log_file(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "a", encoding="utf-8"):
        pass


def action_log_worker():
    while True:
        log_line = action_log_queue.get()
        try:
            with open(ACTION_LOG_FILE, "a", encoding="utf-8") as f:
                f.write(log_line)
        except Exception as e:
            print("action logging failed:", e)
        finally:
            action_log_queue.task_done()


ensure_action_log_file(ACTION_LOG_FILE)

startup_event = {
    "ts": datetime.now(timezone.utc).isoformat(),
    "event": "запись логов началась",
}
with open(ACTION_LOG_FILE, "a", encoding="utf-8") as f:
    f.write(json.dumps(startup_event, ensure_ascii=False) + "\n")

threading.Thread(target=action_log_worker, daemon=True).start()


def pg_query(sql, params=None):
    conn = psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        dbname=PG_DB,
        user=PG_USER,
        password=PG_PASS
    )
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params or {})
            rows = cur.fetchall()
            conn.commit()
            return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/health")
def health():
    try:
        r = requests.get(OPENSEARCH_URL, timeout=TIMEOUT)
        return jsonify({"ok": True, "opensearch_status": r.status_code})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/api/v1/search")
def universal_search():
    q = (request.args.get("q") or "").strip()
    index = (request.args.get("index") or DEFAULT_INDEX).strip()
    size = int(request.args.get("size") or "20")

    if not q:
        return jsonify({"total": 0, "rows": []})

    index_config = {
        "class_tree_v1": {
            "fields": SEARCH_FIELDS,
            "source": ["l4_name", "path_name", "l4_code"]
        },
        "class_tree_nomen_v1": {
            "fields": ["item_name^3", "id"],
            "source": ["item_name", "id"]
        },
        "v_nomenclature_spec_pairs_v1": {
            "fields": ["name_tek^2", "name_tep_korr^2", "id"],
            "source": ["id", "name_tek", "name_tep_korr"]
        }
    }

    cfg = index_config.get(index, {"fields": ["*"], "source": True})

    body = {
        "size": size,
        "_source": cfg["source"],
        "query": {
            "multi_match": {
                "query": q,
                "fields": cfg["fields"],
                "type": "best_fields",
                "fuzziness": "AUTO"
            }
        }
    }

    try:
        url = f"{OPENSEARCH_URL}/{index}/_search"
        r = requests.post(url, json=body, timeout=TIMEOUT)
        r.raise_for_status()

        data = r.json()
        hits_data = data.get("hits", {})
        total = hits_data.get("total", {}).get("value", 0)
        rows = [h["_source"] for h in hits_data.get("hits", [])]

        return jsonify({"total": total, "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e), "total": 0, "rows": []}), 500


@app.post("/v1/actions")
def track_action():
    payload = request.get_json(force=True, silent=True)
    if payload is None:
        return jsonify({"error": "invalid json"}), 400

    event = {
        "ts": datetime.now(timezone.utc).isoformat(),
        "ip": request.headers.get("X-Real-IP") or request.remote_addr,
        "user_agent": request.headers.get("User-Agent", ""),
        "path": request.path,
        "data": payload,
    }

    log_line = json.dumps(event, ensure_ascii=False) + "\n"
    try:
        action_log_queue.put_nowait(log_line)
        return jsonify({"ok": True, "queued": True}), 202
    except Full:
        return jsonify({"ok": False, "queued": False, "error": "log queue is full"}), 503
    
@app.get("/pairs_list")
def pairs_list():
    limit  = int(request.args.get("limit") or "200")
    offset = int(request.args.get("offset") or "0")

    # защита от случайных огромных запросов
    if limit < 1: limit = 1
    if limit > 2000: limit = 2000
    if offset < 0: offset = 0

    sql = """
select id, name_tek, name_tep_korr
from public.v_nomenclature_spec_pairs_v1
order by id
limit %(limit)s offset %(offset)s
    """
    rows = pg_query(sql, {"limit": limit, "offset": offset})

    sql_cnt = "select count(*) as cnt from public.v_nomenclature_spec_pairs_v1"
    cnt = pg_query(sql_cnt)[0]["cnt"]

    return jsonify({"rows": rows, "total": cnt, "limit": limit, "offset": offset})


@app.get("/pairs_rows")
def pairs_rows():
    ids_raw = (request.args.get("ids") or "").strip()
    if not ids_raw:
        return jsonify({"rows": []})

    ids = []
    for x in ids_raw.split(","):
        x = x.strip()
        if x.isdigit():
            ids.append(int(x))

    if not ids:
        return jsonify({"rows": []})

    sql = """
select id, name_tek, name_tep_korr
from public.v_nomenclature_spec_pairs_v1
where id = any(%(ids)s)
order by id
    """
    return jsonify({"rows": pg_query(sql, {"ids": ids})})

@app.post("/reindex")
def reindex():
    script = "/scripts/etl_pg_to_opensearch.py"

    def run():
        try:
            subprocess.run(["python3", script], check=True)
        except Exception as e:
            print("reindex failed:", e)

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "started": True})

@app.post("/reindex_nomen")
def reindex_nomen():
    script = "/scripts/etl_nomen_to_opensearch.py"

    def run():
        try:
            subprocess.run(["python3", script], check=True)
        except Exception as e:
            print("Nomen reindex failed:", e)

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "message": "Nomenclature reindex started"})


@app.patch("/pairs_update/<int:row_id>")
def pairs_update(row_id: int):
    data = request.get_json(force=True, silent=True) or {}
    name_tep_korr = (data.get("name_tep_korr") or "")

    # зеркально пишем name_tek с обрезанием до 150 символов
    name_tek = name_tep_korr[:150]

    sql = """
update public.v_nomenclature_spec_pairs_v1
set name_tep_korr = %(name_tep_korr)s,
    name_tek      = %(name_tek)s
where id = %(id)s
returning id, name_tek, name_tep_korr
    """

    try:
        rows = pg_query(sql, {"id": row_id, "name_tep_korr": name_tep_korr, "name_tek": name_tek})
        if not rows:
            return jsonify({"error": "row not found"}), 404
        return jsonify(rows[0])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.get("/reports/nomenclature_by_object")
def report_nomenclature_by_object():
    sql = """
      select
        o.subobject_name  as object_name,
        count(n.id)   as cnt
      from public.nomenclature_spec_all_v1 n
      left join public.object_v1 o on n.object_id = o.id
      group by o.subobject_name
      order by o.subobject_name
    """
    return jsonify(pg_query(sql))
    
@app.get("/export/vom.xlsx")
def export_vom_xlsx():
    # 1) запрос (ты говорил: l4_name заменил на item_name — значит берём item_name)
    sql = """
    SELECT DISTINCT
      level,
      l1_code, l1_name,
      l2_code, l2_name,
      l3_code, l3_name,
      l4_code,
      item_name,
      l1_num, l2_num, l3_num, l4_num
    FROM public.class_tree_nomen_v1
    WHERE level IN (1,2,3,4)
    ORDER BY l1_num, l2_num, l3_num, l4_num, level, l1_code, l2_code, l3_code, l4_code;
    """
    rows = pg_query(sql)

    # 2) читаем шаблон
    if not os.path.exists(XLSX_TEMPLATE):
        return jsonify({"error": f"template not found: {XLSX_TEMPLATE}"}), 500

    wb = load_workbook(XLSX_TEMPLATE)
    if XLSX_SHEET not in wb.sheetnames:
        return jsonify({"error": f"sheet not found: {XLSX_SHEET}"}), 500

    ws = wb[XLSX_SHEET]
    max_col = ws.max_column

    # строки-образцы стилей в шаблоне
    ROW_L1, ROW_L2, ROW_L3, ROW_L4 = 2, 3, 4, 5

    def snapshot_row_style(src_row: int):
        snap = []
        for col in range(1, max_col + 1):
            c = ws.cell(src_row, col)
            snap.append({
                "style": copy(c._style),
                "number_format": c.number_format,
                "alignment": copy(c.alignment),
                "font": copy(c.font),
                "border": copy(c.border),
                "fill": copy(c.fill),
                "protection": copy(c.protection),
            })
        return snap

    STYLE_L1 = snapshot_row_style(ROW_L1)
    STYLE_L2 = snapshot_row_style(ROW_L2)
    STYLE_L3 = snapshot_row_style(ROW_L3)
    STYLE_L4 = snapshot_row_style(ROW_L4)

    def apply_row_style(dst_row: int, style_snap):
        for col in range(1, max_col + 1):
            dst = ws.cell(dst_row, col)
            s = style_snap[col - 1]
            dst._style = copy(s["style"])
            dst.number_format = s["number_format"]
            dst.alignment = copy(s["alignment"])
            dst.font = copy(s["font"])
            dst.border = copy(s["border"])
            dst.fill = copy(s["fill"])
            dst.protection = copy(s["protection"])

    # 3) чистим лист ниже заголовка
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    out_row = 2

    def write_line(level: int, code: str, name: str, style_snap):
        nonlocal out_row
        ws.insert_rows(out_row)
        apply_row_style(out_row, style_snap)
        ws.cell(out_row, 1).value = level
        ws.cell(out_row, 2).value = code
        ws.cell(out_row, 5).value = name
        out_row += 1

    # 4) пишем как есть (по level)
    for r in rows:
        lvl = int(r.get("level") or 0)

        if lvl == 1:
            code, name, style = r.get("l1_code"), r.get("l1_name"), STYLE_L1
        elif lvl == 2:
            code, name, style = r.get("l2_code"), r.get("l2_name"), STYLE_L2
        elif lvl == 3:
            code, name, style = r.get("l3_code"), r.get("l3_name"), STYLE_L3
        elif lvl == 4:
            code, name, style = r.get("l4_code"), r.get("item_name"), STYLE_L4
        else:
            continue

        write_line(lvl, code, name, style)

    # перенос строк по столбцу E
    for rr in range(1, ws.max_row + 1):
        c = ws.cell(rr, 5)
        a = copy(c.alignment)
        a.wrap_text = True
        c.alignment = a

    # 5) сохраняем во временный файл и отдаём
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()

    wb.save(tmp_path)

    return send_file(
        tmp_path,
        as_attachment=True,
        download_name="vom_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8010)
